import sqlite3
import os
import io
from flask import Flask, jsonify, request, render_template, g, send_file
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import requests
from bs4 import BeautifulSoup
from fpdf import FPDF
import time

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), 'kraduty.db')
CURRENT_YEAR = 2025

# Cache for exchange rates
RATES_CACHE = {
    'timestamp': 0,
    'data': {
        'USD': 129.50, # Fallbacks
        'GBP': 165.00,
        'EUR': 140.00,
        'JPY': 0.85,
        'ZAR': 7.20,
        'AED': 35.20
    }
}

def get_live_rates():
    global RATES_CACHE
    now = time.time()
    # Cache for 6 hours
    if now - RATES_CACHE['timestamp'] < 21600:
        return RATES_CACHE['data']
    
    try:
        url = "https://www.centralbank.go.ke/forex/"
        headers = {'User-Agent': 'Mozilla/5.0'}
        resp = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(resp.content, 'html.parser')
        table = soup.find('table')
        if table:
            new_rates = {}
            rows = table.find_all('tr')[1:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) >= 2:
                    curr = cols[0].text.strip().upper()
                    try:
                        rate = float(cols[1].text.strip())
                        if 'US DOLLAR' in curr: new_rates['USD'] = rate
                        elif 'STG POUND' in curr: new_rates['GBP'] = rate
                        elif 'EURO' in curr: new_rates['EUR'] = rate
                        elif 'JPY (100)' in curr: new_rates['JPY'] = rate / 100.0
                        elif 'AE DIRHAM' in curr: new_rates['AED'] = rate
                        elif 'SA RAND' in curr: new_rates['ZAR'] = rate
                    except: continue
            if new_rates:
                RATES_CACHE['data'].update(new_rates)
                RATES_CACHE['timestamp'] = now
    except Exception as e:
        print(f"Rate fetch error: {e}")
    
    return RATES_CACHE['data']

class DutyPDF(FPDF):
    def header(self):
        self.set_font('Helvetica', 'B', 15)
        self.set_text_color(27, 77, 56) # Professional Green
        self.cell(0, 10, 'KRA MOTOR VEHICLE DUTY QUOTE', align='C', new_x='LMARGIN', new_y='NEXT')
        self.set_font('Helvetica', 'I', 9)
        self.set_text_color(100)
        self.cell(0, 5, f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M")}', align='C', new_x='LMARGIN', new_y='NEXT')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.cell(0, 10, 'Page ' + str(self.page_no()) + ' - KRA Duty Calculator', align='C')

@app.route('/api/exchange-rates')
def api_exchange_rates():
    return jsonify(get_live_rates())

@app.route('/api/report/pdf-quote', methods=['POST'])
def api_pdf_quote():
    data = request.get_json()
    
    pdf = DutyPDF()
    pdf.add_page()
    
    # Vehicle Info Section
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 10, ' VEHICLE SPECIFICATIONS', fill=True, new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 10)
    
    info = [
        ('Make/Model:', f"{data.get('make')} {data.get('model')}"),
        ('Year of Manufacture:', str(data.get('yom'))),
        ('Engine Capacity:', str(data.get('engine_capacity'))),
        ('Fuel Type:', str(data.get('fuel'))),
        ('Body Type:', str(data.get('body_type'))),
        ('CRSP Value:', f"KES {int(data.get('crsp', 0)):,}")
    ]
    
    for label, val in info:
        pdf.cell(50, 8, label)
        pdf.cell(0, 8, val, new_x='LMARGIN', new_y='NEXT')
    
    pdf.ln(5)
    
    # Duty Breakdown Section
    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 10, ' DUTY & TAX BREAKDOWN', fill=True, new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 10)
    
    taxes = [
        ('Customs Value', data.get('customs_value')),
        (f"Import Duty ({data.get('import_duty_rate')})", data.get('import_duty')),
        (f"Excise Duty ({data.get('excise_rate')})", data.get('excise_duty')),
        ('VAT (16%)', data.get('vat')),
        ('RDL (2%)', data.get('rdl')),
        ('IDF Fees (2.5%)', data.get('idf'))
    ]
    
    for label, val in taxes:
        pdf.cell(100, 8, label)
        pdf.cell(0, 8, f"KES {int(val):,}", align='R', new_x='LMARGIN', new_y='NEXT')
        
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_text_color(27, 77, 56)
    pdf.cell(100, 10, 'TOTAL DUTY PAYABLE', border='T')
    pdf.cell(0, 10, f"KES {int(data.get('grand_total', 0)):,}", border='T', align='R', new_x='LMARGIN', new_y='NEXT')
    pdf.set_text_color(0)
    
    # Landing Cost Section if present
    if data.get('landed_cost'):
        pdf.ln(5)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 10, ' TOTAL LANDED COST ESTIMATE', fill=True, new_x='LMARGIN', new_y='NEXT')
        pdf.set_font('Helvetica', '', 10)
        
        pdf.cell(100, 8, 'Purchase + Shipping (Converted)')
        pdf.cell(0, 8, f"KES {int(data.get('fob_kes', 0)):,}", align='R', new_x='LMARGIN', new_y='NEXT')
        pdf.cell(100, 8, 'Total Duty Payable')
        pdf.cell(0, 8, f"KES {int(data.get('grand_total', 0)):,}", align='R', new_x='LMARGIN', new_y='NEXT')
        
        pdf.set_font('Helvetica', 'B', 11)
        pdf.set_text_color(180, 50, 50) # Red-ish for bottom line
        pdf.cell(100, 10, 'ESTIMATED TOTAL LANDED COST', border='T')
        pdf.cell(0, 10, f"KES {int(data.get('landed_cost', 0)):,}", border='T', align='R', new_x='LMARGIN', new_y='NEXT')
        
    pdf.ln(10)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.multi_cell(0, 5, 'Disclaimer: This quote is an estimate based on current KRA CRSP values and prevailing exchange rates. Actual values at the time of entry may vary. This is not an official KRA assessment.')

    output = io.BytesIO()
    pdf_bytes = pdf.output()
    output.write(pdf_bytes)
    output.seek(0)
    
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name='KRA_Duty_Quote.pdf')

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()

# Depreciation tables
DIRECT_DEPR = [
    (0, 2, 0.20),  # 1 <=2 years
    (2, 3, 0.30),  # 2 <=3 years
    (3, 4, 0.40),  # 3 <=4 years
    (4, 5, 0.50),  # 4 <=5 years
    (5, 6, 0.55),  # 5 <=6 years
    (6, 7, 0.60),  # 6 <=7 years
    (7, 8, 0.65),  # 7 <=8 years
]

PREV_REG_DEPR = [
    (0, 1, 0.20),
    (1, 2, 0.35),
    (2, 3, 0.50),
    (3, 4, 0.60),
    (4, 5, 0.70),
    (5, 6, 0.75),
    (6, 7, 0.80),
    (7, 8, 0.83),
    (8, 9, 0.86),
    (9, 10, 0.89),
    (10, 11, 0.90),
    (11, 12, 0.91),
    (12, 13, 0.92),
    (13, 14, 0.93),
    (14, 15, 0.94),
] # over 15 years 95%

def get_depreciation(years_old, is_direct_import=True):
    if is_direct_import:
        for lo, hi, rate in DIRECT_DEPR:
            if lo < years_old <= hi:
                return rate
        if years_old > 8: return 0.70 # Default cap for older if allowed
        return 0.0
    else:
        for lo, hi, rate in PREV_REG_DEPR:
            if lo < years_old <= hi:
                return rate
        if years_old > 15: return 0.95
        return 0.0

def calc_taxes(crsp, years_old, is_direct, vehicle_type, engine_cc, fuel, body_type=None):
    dep_rate = get_depreciation(years_old, is_direct)
    
    # Extra Depreciation - Defaulting to 0 for now as it's not in UI
    extra_dep = 0.0
    
    fuel_upper = fuel.upper() if fuel else ''
    body_upper = body_type.upper() if body_type else ''
    
    # Clean CC
    try:
        cc_str = str(engine_cc).split('(')[0].split(' kWh')[0].split(' ')[0] if engine_cc else '0'
        cc = float(cc_str) if cc_str.replace('.', '', 1).isdigit() else 0
    except (ValueError, TypeError):
        cc = 0

    # 1. Determine Tabulation
    tab = 2 # Default to Tab 2
    
    eng_upper = str(engine_cc).upper() if engine_cc else ''
    
    # Improved is_electric detection
    # A car is electric if fuel says ELECTRIC OR engine_capacity says EV/KWH/HP and it's NOT a hybrid
    is_hybrid = 'HYBRID' in fuel_upper or 'EREV' in eng_upper or 'PHEV' in eng_upper
    is_electric = ('ELECTRIC' in fuel_upper or 'EV' in eng_upper or 'KWH' in eng_upper or ' HP' in eng_upper) and not is_hybrid
    
    if vehicle_type == 'motor_cycle':
        tab = 9
    elif vehicle_type == 'tractor' or body_upper in ('TRACTOR', 'HEAVY MACHINERY', 'GRADER', 'MIXER', 'TRANSIT MIXER'):
        tab = 10
    elif 'AMBULANCE' in body_upper:
        tab = 8
    elif body_upper in ('PRIME MOVER', 'PM', 'PRIM£ MOVER'):
        tab = 6
    elif 'TRAILER' in body_upper:
        tab = 7
    elif is_electric:
        tab = 4
    elif 'SCHOOL BUS' in body_upper:
        tab = 5
    elif fuel_upper in ('GASOLINE', 'PETROL') and cc > 3000:
        tab = 3
    elif fuel_upper == 'DIESEL' and cc > 2500:
        tab = 3
    elif cc <= 1500:
        tab = 1
    else:
        tab = 2

    # 2. Assign Rates and Divisors based on Tab
    import_rate = 0.35
    excise_rate = 0.25
    vat_rate = 0.16
    
    # Divisors: d1=Import, d2=Excise, d3=VAT
    d1, d2, d3 = 1.35, 1.25, 1.16
    
    if tab == 1:
        import_rate, excise_rate = 0.35, 0.20
        d1, d2, d3 = 1.35, 1.20, 1.16
    elif tab == 2:
        import_rate, excise_rate = 0.35, 0.25
        d1, d2, d3 = 1.35, 1.25, 1.16
    elif tab == 3:
        import_rate, excise_rate = 0.35, 0.35
        d1, d2, d3 = 1.35, 1.35, 1.16
    elif tab == 4:
        import_rate, excise_rate = 0.25, 0.10
        d1, d2, d3 = 1.25, 1.10, 1.16
    elif tab == 5:
        import_rate, excise_rate = 0.35, 0.25
        d1, d2, d3 = 1.35, 1.25, 1.16
    elif tab == 6 or tab == 7:
        import_rate, excise_rate = 0.35, 0.00
        d1, d2, d3 = 1.35, 1.0, 1.16
    elif tab == 8:
        import_rate, excise_rate = 0.00, 0.25
        d1, d2, d3 = 1.25, 1.16, 1.0 # Formula says /1.25/1.16. 1.25 is for Excise.
        # Wait, Tab 8 formula: ((CRSP/1.25)*(100%-Depreciation)/1.25/1.16)
        # Import Duty 0%, Excise Duty 25%, VAT 16%
        # So it's CRSP/1.25 * (1-dep) / 1.25 (excise) / 1.16 (vat)
        d1, d2, d3 = 1.0, 1.25, 1.16
    elif tab == 9:
        import_rate = 0.25
        excise_flat = 12952.83
        d1, d2, d3 = 1.25, 1.0, 1.16
    elif tab == 10:
        import_rate, excise_rate = 0.00, 0.00
        d1, d2, d3 = 1.0, 1.0, 1.16

    # 3. Calculate Customs Value
    # Custom Value = ((CRSP/1.25)*(100%-Depreciation)/divisor1/divisor2/divisor3)*(100%-Extra Depreciation)
    if tab == 8: # Special case divisor handling based on user prompt
        customs = ((crsp / 1.25) * (1 - dep_rate) / 1.25 / 1.16) * (1 - extra_dep)
    elif tab == 9:
        customs = ((crsp / 1.25) * (1 - dep_rate) / 1.25 / 1.16) * (1 - extra_dep)
    elif tab == 10:
        customs = ((crsp / 1.25) * (1 - dep_rate) / 1.16) * (1 - extra_dep)
    elif tab == 6 or tab == 7:
        customs = ((crsp / 1.25) * (1 - dep_rate) / 1.35 / 1.16) * (1 - extra_dep)
    else:
        customs = ((crsp / 1.25) * (1 - dep_rate) / d1 / d2 / d3) * (1 - extra_dep)

    # 4. Calculate Duties
    import_duty = customs * import_rate
    
    if tab == 9:
        excise_duty = excise_flat
    else:
        excise_val = customs + import_duty
        excise_duty = excise_val * excise_rate
        
    vat_val = customs + import_duty + excise_duty
    vat = vat_val * vat_rate
    
    # RDL and IDF (Exempt for previously registered)
    if is_direct:
        rdl = customs * 0.02
        idf = customs * 0.025
    else:
        rdl = 0.0
        idf = 0.0
        
    grand_total = import_duty + excise_duty + vat + rdl + idf

    # Basis Description
    basis_map = {
        1: "Engine capacity \u2264 1500cc (including S/Cab, Lorry, Bus)",
        2: "Engine rating > 1500cc (Standard Passenger/Cargo)",
        3: f"High capacity engine ({cc}cc) \u2014 exceeding limits for {'Petrol' if fuel_upper in ('GASOLINE', 'PETROL') else 'Diesel'}",
        4: "100% Electric Powered Vehicle",
        5: "Public School Bus",
        6: "Prime Mover (Excise Exempt)",
        7: "Trailer (Excise Exempt)",
        8: "Ambulance (Duty Exempt)",
        9: "Motor Cycle (Flat Excise)",
        10: "Heavy Machinery / Tractor (Duty & Excise Exempt)"
    }
    tax_basis = basis_map.get(tab, "Standard Valuation")

    components = {
        'customs_value': round(customs, 2),
        'import_duty': round(import_duty, 2),
        'import_duty_rate': f'{int(import_rate*100)}%',
        'excise_duty': round(excise_duty, 2),
        'excise_rate': f'{int(excise_rate*100)}%' if tab != 9 else 'Flat KES 12,952.83',
        'vat_value': round(vat_val, 2),
        'vat': round(vat, 2),
        'rdl': round(rdl, 2),
        'idf': round(idf, 2),
        'grand_total': round(grand_total, 2),
        'tabulation': tab,
        'tax_basis': tax_basis
    }
    return components

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/makes')
def api_makes():
    db = get_db()
    cur = db.execute("SELECT DISTINCT make FROM motor_vehicles WHERE make IS NOT NULL ORDER BY make")
    makes = [row['make'] for row in cur.fetchall()]
    return jsonify(makes)

@app.route('/api/models')
def api_models():
    make = request.args.get('make', '')
    db = get_db()
    cur = db.execute("SELECT DISTINCT model FROM motor_vehicles WHERE make = ? AND model IS NOT NULL ORDER BY model", (make,))
    models = [row['model'] for row in cur.fetchall()]
    return jsonify(models)

@app.route('/api/model_numbers')
def api_model_numbers():
    make = request.args.get('make', '')
    model = request.args.get('model', '')
    db = get_db()
    cur = db.execute(
        "SELECT DISTINCT model_number FROM motor_vehicles WHERE make = ? AND model = ? AND model_number IS NOT NULL ORDER BY model_number",
        (make, model))
    nums = [row['model_number'] for row in cur.fetchall()]
    return jsonify(nums)

@app.route('/api/variants')
def api_variants():
    make = request.args.get('make', '')
    model = request.args.get('model', '')
    model_no = request.args.get('model_number', '')
    db = get_db()
    if model_no:
        cur = db.execute(
            "SELECT DISTINCT transmission, drive_config, engine_capacity, body_type, fuel, crsp FROM motor_vehicles WHERE make = ? AND model = ? AND model_number = ?",
            (make, model, model_no))
    else:
        cur = db.execute(
            "SELECT DISTINCT transmission, drive_config, engine_capacity, body_type, fuel, crsp FROM motor_vehicles WHERE make = ? AND model = ?",
            (make, model))
    variants = []
    for row in cur.fetchall():
        variants.append(dict(zip(['transmission', 'drive_config', 'engine_capacity', 'body_type', 'fuel', 'crsp'], row)))
    return jsonify(variants)

@app.route('/api/mc_makes')
def api_mc_makes():
    db = get_db()
    cur = db.execute("SELECT DISTINCT make FROM motor_cycles WHERE make IS NOT NULL ORDER BY make")
    return jsonify([row['make'] for row in cur.fetchall()])

@app.route('/api/mc_models')
def api_mc_models():
    make = request.args.get('make', '')
    db = get_db()
    cur = db.execute("SELECT DISTINCT model FROM motor_cycles WHERE make = ? AND model IS NOT NULL ORDER BY model", (make,))
    return jsonify([row['model'] for row in cur.fetchall()])

@app.route('/api/mc_model_numbers')
def api_mc_model_numbers():
    make = request.args.get('make', '')
    model = request.args.get('model', '')
    db = get_db()
    cur = db.execute(
        "SELECT DISTINCT model_number FROM motor_cycles WHERE make = ? AND model = ? AND model_number IS NOT NULL ORDER BY model_number",
        (make, model))
    return jsonify([row['model_number'] for row in cur.fetchall()])

@app.route('/api/mc_variants')
def api_mc_variants():
    make = request.args.get('make', '')
    model = request.args.get('model', '')
    model_no = request.args.get('model_number', '')
    db = get_db()
    if model_no:
        cur = db.execute(
            "SELECT DISTINCT transmission, engine_capacity, fuel, crsp FROM motor_cycles WHERE make = ? AND model = ? AND model_number = ?",
            (make, model, model_no))
    else:
        cur = db.execute(
            "SELECT DISTINCT transmission, engine_capacity, fuel, crsp FROM motor_cycles WHERE make = ? AND model = ?",
            (make, model))
    variants = []
    for row in cur.fetchall():
        v = dict(zip(['transmission', 'engine_capacity', 'fuel', 'crsp'], row))
        if v['engine_capacity'] is not None:
            v['engine_capacity'] = str(int(v['engine_capacity']))
        v['drive_config'] = None
        v['body_type'] = None
        variants.append(v)
    return jsonify(variants)

@app.route('/api/tractor_makes')
def api_tractor_makes():
    db = get_db()
    cur = db.execute("SELECT DISTINCT make FROM tractors WHERE make IS NOT NULL ORDER BY make")
    return jsonify([row['make'] for row in cur.fetchall()])

@app.route('/api/tractor_models')
def api_tractor_models():
    make = request.args.get('make', '')
    db = get_db()
    cur = db.execute("SELECT model, horsepower, crsp FROM tractors WHERE make = ? AND model IS NOT NULL ORDER BY model", (make,))
    models = []
    for row in cur.fetchall():
        models.append({'model': row['model'], 'horsepower': row['horsepower'], 'crsp': row['crsp']})
    return jsonify(models)

@app.route('/api/body_types')
def api_body_types():
    db = get_db()
    cur = db.execute("SELECT DISTINCT body_type FROM motor_vehicles WHERE body_type IS NOT NULL ORDER BY body_type")
    return jsonify([row['body_type'] for row in cur.fetchall()])

@app.route('/api/calculate', methods=['POST'])
def api_calculate():
    data = request.get_json()
    crsp = float(data.get('crsp', 0))
    year_of_manufacture = int(data.get('yom', CURRENT_YEAR))
    month_of_manufacture = int(data.get('mom', 1))
    is_direct = data.get('is_direct', True)
    vehicle_type = data.get('vehicle_type', 'motor_vehicle')
    engine_cc = data.get('engine_capacity', '0')
    fuel = data.get('fuel', 'GASOLINE')
    body_type = data.get('body_type', '')

    if isinstance(is_direct, str):
        is_direct = is_direct == 'true'

    calendar_age = CURRENT_YEAR - year_of_manufacture
    if vehicle_type == 'motor_vehicle' and is_direct and calendar_age > 7:
        return jsonify({'error': 'Vehicles older than 7 years (manufactured before 2018) cannot be imported into Kenya.'}), 400

    current_month = 6
    total_months = (CURRENT_YEAR * 12 + current_month) - (year_of_manufacture * 12 + month_of_manufacture)
    age_years = total_months / 12.0

    result = calc_taxes(crsp, age_years, is_direct, vehicle_type, engine_cc, fuel, body_type)
    result['crsp'] = crsp
    result['yom'] = year_of_manufacture
    result['mom'] = month_of_manufacture
    result['age_years'] = round(age_years, 1)
    result['depreciation_rate'] = f'{get_depreciation(age_years, is_direct) * 100:.0f}%'

    return jsonify(result)

from openpyxl.utils import get_column_letter

@app.route('/api/report/duties-below')
def api_report_duties_below():
    yom = int(request.args.get('yom', CURRENT_YEAR))
    mom = int(request.args.get('mom', 1))
    max_duty = float(request.args.get('max_duty', 500000))
    is_direct = request.args.get('is_direct', 'true') == 'true'
    vehicle_type = request.args.get('vehicle_type', 'motor_vehicle')
    body_filter = request.args.get('body_type', '')

    db = get_db()

    if vehicle_type == 'motor_vehicle':
        if body_filter:
            cur = db.execute("SELECT * FROM motor_vehicles WHERE body_type = ?", (body_filter,))
        else:
            cur = db.execute("SELECT * FROM motor_vehicles")
        rows = cur.fetchall()
    elif vehicle_type == 'motor_cycle':
        cur = db.execute("SELECT * FROM motor_cycles")
        rows = cur.fetchall()
    else:
        return jsonify({'error': 'Unsupported vehicle type'}), 400

    current_month = 6
    total_months = (CURRENT_YEAR * 12 + current_month) - (yom * 12 + mom)
    age_years = total_months / 12.0
    dep_rate = get_depreciation(age_years, is_direct)

    results = []
    for row in rows:
        row = dict(row)
        crsp = row.get('crsp')
        if not crsp:
            continue

        if vehicle_type == 'motor_vehicle' and is_direct and (CURRENT_YEAR - yom) > 7:
            continue

        engine_cc = str(row.get('engine_capacity') or '0')
        fuel = str(row.get('fuel') or 'GASOLINE')
        body_type = str(row.get('body_type') or '')

        tax = calc_taxes(crsp, age_years, is_direct, vehicle_type, engine_cc, fuel, body_type)

        if tax['grand_total'] < max_duty:
            entry = {**row, **tax}
            entry['age_years'] = round(age_years, 1)
            entry['depreciation_rate'] = f'{dep_rate * 100:.0f}%'
            results.append(entry)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duties Below Threshold"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4D38", end_color="1B4D38", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    even_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'), 
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), 
        bottom=Side(style='thin', color='DDDDDD')
    )

    if vehicle_type == 'motor_vehicle':
        cols = [
            'make', 'model', 'model_number', 'transmission', 'drive_config',
            'engine_capacity', 'body_type', 'gvw', 'seating', 'fuel', 'crsp',
            'age_years', 'depreciation_rate',
            'grand_total', 'customs_value', 'import_duty', 'excise_duty',
            'vat', 'rdl', 'idf'
        ]
        headers = [
            'Make', 'Model', 'Model Number', 'Transmission', 'Drive Config',
            'Engine Capacity', 'Body Type', 'GVW', 'Seating', 'Fuel', 'CRSP (KES)',
            'Age (yrs)', 'Depreciation Rate',
            'Total Duty (KES)', 'Customs Value (KES)', 'Import Duty (KES)',
            'Excise Duty (KES)', 'VAT (KES)', 'RDL (KES)', 'IDF (KES)'
        ]
    else:
        cols = [
            'make', 'model', 'model_number', 'transmission',
            'engine_capacity', 'seating', 'fuel', 'crsp',
            'age_years', 'depreciation_rate',
            'grand_total', 'customs_value', 'import_duty', 'excise_duty',
            'vat', 'rdl', 'idf'
        ]
        headers = [
            'Make', 'Model', 'Model Number', 'Transmission',
            'Engine Capacity', 'Seating', 'Fuel', 'CRSP (KES)',
            'Age (yrs)', 'Depreciation Rate',
            'Total Duty (KES)', 'Customs Value (KES)', 'Import Duty (KES)',
            'Excise Duty (KES)', 'VAT (KES)', 'RDL (KES)', 'IDF (KES)'
        ]

    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write Data
    for row_idx, entry in enumerate(results, 2):
        for col_idx, key in enumerate(cols, 1):
            val = entry.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            
            # Alternating rows
            if row_idx % 2 == 0:
                cell.fill = even_fill

            # Format numbers
            if isinstance(val, (int, float)):
                if key in ('crsp', 'grand_total', 'customs_value', 'import_duty', 'excise_duty', 'vat', 'rdl', 'idf'):
                    cell.value = round(val)
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Add Filter
    ws.auto_filter.ref = ws.dimensions

    # Adjust Column Widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(13, len(header) + 2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'duties_below_{int(max_duty)}_{vehicle_type}_yom{yom}_mom{mom}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5812))
    app.run(host='0.0.0.0', port=port, debug=False)
