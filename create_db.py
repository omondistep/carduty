import sqlite3
import openpyxl
import os

DB_PATH = os.path.join(os.path.dirname(__file__), 'kraduty.db')
XLSX_PATH = os.path.join(os.path.dirname(__file__), 'New-CRSP---July-2025.xlsx')

def safe_float(v):
    if v is None:
        return None
    try:
        return float(str(v).split('(')[0].strip().replace(',', ''))
    except (ValueError, TypeError):
        return None

def safe_int(v):
    if v is None:
        return None
    try:
        return int(float(str(v).split('(')[0].split('–')[0].strip()))
    except (ValueError, TypeError):
        digits = ''.join(c for c in str(v) if c.isdigit())
        return int(digits[0]) if digits else None

import re

def clean(s):
    if s is None: return None
    s = str(s).strip()
    return re.sub(r'\s+', ' ', s) if s else None

def normalize_general(s):
    c = clean(s)
    return c.upper() if c else None

def normalize_model(s):
    val = normalize_general(s)
    if not val: return None
    # Remove spaces around certain characters to unify e.g. "D: 5" and "D:5"
    val = re.sub(r'\s*([+:\-/])\s*', r'\1', val)
    # Handle specific common cases like "TOWN ACE" vs "TOWNACE"
    val = val.replace('TOWN ACE', 'TOWNACE')
    val = val.replace('LAND MARK', 'LANDMARK')
    return val

def normalize_transmission(s):
    val = normalize_general(s)
    if not val: return None
    # Remove all spaces for codes like "6 MT" -> "6MT"
    val = "".join(val.split())
    if "MANUAL" in val: return "MANUAL"
    return val

def normalize_drive(s):
    val = normalize_general(s)
    if not val: return None
    val = val.replace('*', 'X').replace('×', 'X')
    val = "".join(val.split())
    # Handle common typos
    if val == "4WDD": return "4WD"
    return val

def normalize_body(s):
    val = normalize_general(s)
    if not val: return None
    
    # Mappings
    if val in ('SAL', 'SALOON', 'SEDAN'): return 'SALOON'
    if val in ('HATCBACK', 'HATCHBACK'): return 'HATCHBACK'
    if val in ('S/WAGON', 'S. WAGON', 'STATION WAGON', 'WAGON'): return 'STATION WAGON'
    if val in ('D/CAB', 'DOUBLE CAB', 'DOUBLE CABIN', 'DUAL CAB', 'CREW CAB'): return 'DOUBLE CAB'
    if val in ('S/CAB', 'S/CABIN', 'SINGLE CAB', 'SINGLE CABIN'): return 'SINGLE CAB'
    if val in ('PICK UP', 'PICKUP'): return 'PICKUP'
    if val in ('TRK', 'TRUCK'): return 'TRUCK'
    if val in ('PRIM£ MOVER', 'PM', 'PRIME MOVER'): return 'PRIME MOVER'
    if val in ('MINVAN', 'MINIVAN'): return 'MINIVAN'
    if val in ('TRANSIT MIXER', 'MIXER'): return 'TRANSIT MIXER'
    if val == 'CONVRTIBLE': return 'CONVERTIBLE'
    
    return val

def normalize_fuel(fuel):
    if not fuel:
        return None
    cleaned = str(fuel).strip()
    f_upper = "".join(cleaned.upper().split())
    
    if f_upper in ("DIESEL", "DEISEL") or "DEISEL" in f_upper:
        return "DIESEL"
    if f_upper in ("ELECCTRIC", "ELECTRIC", "ELECTRIC(EV)", "ELECTIRC"):
        return "ELECTRIC"
    if f_upper in ("GASOLINE", "GASOLIN"):
        return "GASOLINE"
    if f_upper in ("PETROL", "PETORL"):
        return "PETROL"
    if "HYBRID" in f_upper:
        if "PLUG" in f_upper:
            return "PLUG-IN HYBRID"
        if "PETROL" in f_upper:
            return "PETROL/ELECTRIC"
        return "HYBRID"
    # Filter out numeric junk if any
    if f_upper.isdigit(): return None
    return cleaned.upper()

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

if os.path.exists(DB_PATH):
    os.remove(DB_PATH)

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

cur.execute("""
CREATE TABLE motor_vehicles (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    make TEXT,
    model TEXT,
    model_number TEXT,
    transmission TEXT,
    drive_config TEXT,
    engine_capacity TEXT,
    body_type TEXT,
    gvw REAL,
    seating INTEGER,
    fuel TEXT,
    crsp REAL
)
""")

ws = wb['M.Vehicle CRSP July 2025']
count = 0
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    make, model, model_no, trans, drive, eng, body, gvw, seat, fuel, crsp = row[:11]
    if make is not None:
        cur.execute("""INSERT INTO motor_vehicles 
            (make, model, model_number, transmission, drive_config, engine_capacity, body_type, gvw, seating, fuel, crsp)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (normalize_general(make), normalize_model(model), normalize_general(model_no), 
             normalize_transmission(trans), normalize_drive(drive), clean(eng), 
             normalize_body(body), safe_float(gvw), safe_int(seat), normalize_fuel(fuel), safe_float(crsp)))
        count += 1
print(f"Inserted {count} motor vehicles")

cur.execute("""
CREATE TABLE motor_cycles (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    make TEXT,
    model TEXT,
    model_number TEXT,
    transmission TEXT,
    engine_capacity REAL,
    seating INTEGER,
    fuel TEXT,
    crsp REAL
)
""")

ws2 = wb['Motor Cycles July 2025']
count2 = 0
for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, values_only=True):
    make, model, model_no, trans, eng, seat, fuel, crsp = row[:8]
    if make is not None:
        cur.execute("""INSERT INTO motor_cycles
            (make, model, model_number, transmission, engine_capacity, seating, fuel, crsp)
            VALUES (?,?,?,?,?,?,?,?)""",
            (normalize_general(make), normalize_model(model), normalize_general(model_no), 
             normalize_transmission(trans), safe_float(eng), safe_int(seat), normalize_fuel(fuel), safe_float(crsp)))
        count2 += 1
print(f"Inserted {count2} motor cycles")

cur.execute("""
CREATE TABLE tractors (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    make TEXT,
    model TEXT,
    horsepower REAL,
    crsp REAL
)
""")

ws3 = wb['Tractors & Graders July 2025']
count3 = 0
current_make = None
for row in ws3.iter_rows(min_row=3, max_row=ws3.max_row, values_only=True):
    model_val, hp_val, crsp_val = row[0], row[1], row[2]
    if model_val is not None:
        model_str = str(model_val).strip()
        if crsp_val is None and hp_val is None and model_str.isupper():
            current_make = model_str
            continue
        if model_str == 'KSHS':
            continue
    if current_make and model_val and crsp_val:
        cur.execute("""INSERT INTO tractors (make, model, horsepower, crsp) VALUES (?,?,?,?)""",
            (normalize_general(current_make), normalize_model(model_val), safe_float(hp_val), safe_float(crsp_val)))
        count3 += 1
print(f"Inserted {count3} tractors")

conn.commit()
conn.close()
print(f"Database created at {DB_PATH}")
